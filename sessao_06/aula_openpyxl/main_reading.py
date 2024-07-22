from openpyxl import Workbook, load_workbook
from pathlib import Path

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# carregando arquivo do excel
workbook = load_workbook(WORKBOOK_PATH)

sheet_name = 'Estudantes'
worksheet = workbook[sheet_name ]

for row in worksheet.iter_rows(min_row=2):
    for col  in row:
        print(col.value, end='\t')
        if col.value == 'Mario':
            worksheet.cell(col.row, 2, '15')
    print()

print(worksheet['B3'].value)

worksheet['B3'].value = 18

workbook.save(WORKBOOK_PATH)