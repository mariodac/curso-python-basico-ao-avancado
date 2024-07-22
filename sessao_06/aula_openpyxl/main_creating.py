# openpyxl para trabalhar com arquivos Excel xlsx, xlsm, xltx e xltm
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

from openpyxl import Workbook
from pathlib import Path

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
# cria página de planilha padrão 
# worksheet: Worksheet = workbook.active # type: ignore

# cria nome personalizado para página de planilha
sheet_name = 'Estudantes'
# cria a página de planilha na primeira posição
workbook.create_sheet(sheet_name, 0)
# seleciona a planilha
worksheet = workbook[sheet_name ]

# Remover página de planilha
workbook.remove(workbook['Sheet'])

# Criando cabeçalhos
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

students = [
    # nome   idade  nota
    ['João',   14,  5.5],
    ['Maria',  12,  7.0],
    ['Pedro',  15,  8.5],
    ['Ana',    13,  6.8],
]

for i, student_row in enumerate(students, start=2):
    for j, student_col in enumerate(student_row, start=1):
        worksheet.cell(i, j, student_col)

students2 = [
    # nome   idade  nota
    ['José',   16,  9.0],
    ['Mario',  11,  7.5],
    ['Luís',  14,  8.0],
    ['Anais',    12,  6.7],
]

for student in students2:
    # adiciona os dados na próxima linha
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)